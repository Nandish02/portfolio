"""Export ScoredJob list to xlsx with the locked 29-column schema.

Features:
- Frozen header
- Hyperlink on Apply Link
- Data validation dropdown on Status
- Conditional color on MatchScore + Sponsorship + days_open
- Autofilter on the full range
- Sort by MatchScore desc (then by first_seen desc)
"""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..models import ScoredJob


logger = logging.getLogger(__name__)


COLUMNS = [
    ("id", 12),
    ("first_seen", 12),
    ("last_seen", 12),
    ("days_open", 10),
    ("Company", 20),
    ("Title", 40),
    ("Level", 12),
    ("Locations", 30),
    ("Posted", 12),
    ("Deadline", 12),
    ("Salary", 18),
    ("YOE Required", 14),
    ("Sponsorship", 14),
    ("Required Skills", 35),
    ("Nice-to-Have", 30),
    ("Your Skill Gap", 30),
    ("Your Skill Match", 30),
    ("MatchScore", 12),
    ("Fit Blurb", 50),
    ("Tailored Resume Bullets", 60),
    ("Cover Letter Hook", 50),
    ("Job Description", 60),
    ("Apply Link", 50),
    ("Source", 18),
    ("Status", 14),
    ("Date Applied", 12),
    ("Notes", 30),
    ("Referral Available", 14),
    ("Follow-Up Date", 14),
]


_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _detect_level(title: str) -> str:
    t = (title or "").lower()
    if "intern" in t or "co-op" in t or "coop" in t:
        return "Intern"
    if "new grad" in t or "new-grad" in t or "newgrad" in t:
        return "New Grad"
    if "staff" in t:
        return "Staff"
    if "principal" in t:
        return "Principal"
    if "senior" in t or " sr " in f" {t} ":
        return "Senior"
    return "Other"


def _date_or_blank(dt) -> str:
    if not dt:
        return ""
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt)
        except Exception:
            return dt[:10]
    return dt.strftime("%Y-%m-%d")


def _days_open(first_seen) -> int | str:
    if not first_seen:
        return ""
    if isinstance(first_seen, str):
        try:
            first_seen = datetime.fromisoformat(first_seen)
        except Exception:
            return ""
    now = datetime.now(timezone.utc)
    if first_seen.tzinfo is None:
        first_seen = first_seen.replace(tzinfo=timezone.utc)
    return (now - first_seen).days


def export_xlsx(scored: list[ScoredJob], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "internships"

    # header
    for i, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width

    # sort: score desc (None last), then first_seen desc
    def sort_key(sj: ScoredJob):
        sc = sj.score.score if sj.score else -1
        fs = sj.posting.first_seen or datetime.min.replace(tzinfo=timezone.utc)
        return (-sc, -fs.timestamp())

    scored_sorted = sorted(scored, key=sort_key)

    for row_i, sj in enumerate(scored_sorted, start=2):
        jp = sj.posting
        sc = sj.score
        row = [
            jp.canonical_key,
            _date_or_blank(jp.first_seen),
            _date_or_blank(jp.last_seen),
            _days_open(jp.first_seen),
            jp.company,
            jp.title,
            _detect_level(jp.title),
            "; ".join(jp.locations[:6]),
            _date_or_blank(jp.posted_at),
            _date_or_blank(jp.deadline_at),
            jp.salary_text,
            (sc.yoe_required if sc else ""),
            (sc.sponsorship if sc else "Unknown"),
            ", ".join(sc.required_skills) if sc else "",
            ", ".join(sc.nice_to_have) if sc else "",
            ", ".join(sc.your_skill_gap) if sc else "",
            ", ".join(sc.your_skill_match) if sc else "",
            sc.score if sc else "",
            sc.fit_blurb if sc else "",
            "\n".join(f"• {b}" for b in sc.tailored_bullets) if sc else "",
            sc.cover_letter_hook if sc else "",
            (jp.description or "")[:1500],
            jp.apply_url,
            jp.source,
            "New",
            "",       # Date Applied
            "",       # Notes
            "No",     # Referral Available
            "",       # Follow-Up Date
        ]
        for col_i, val in enumerate(row, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        # hyperlink on Apply Link col (col 23)
        if jp.apply_url:
            link_cell = ws.cell(row=row_i, column=23)
            link_cell.hyperlink = jp.apply_url
            link_cell.font = Font(color="0563C1", underline="single")

    n_rows = len(scored_sorted)
    last_col = get_column_letter(len(COLUMNS))
    last_row = max(2, n_rows + 1)

    # freeze header
    ws.freeze_panes = "A2"
    # autofilter
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # conditional formatting
    # MatchScore col (18) — color scale 1=red, 5=yellow, 10=green
    score_range = f"R2:R{last_row}"
    ws.conditional_formatting.add(
        score_range,
        ColorScaleRule(
            start_type="num", start_value=1, start_color="F8696B",
            mid_type="num", mid_value=6, mid_color="FFEB84",
            end_type="num", end_value=10, end_color="63BE7B",
        ),
    )
    # Sponsorship col (13) — red if 'No', green if 'Yes' or 'Likely Yes'
    sp_range = f"M2:M{last_row}"
    ws.conditional_formatting.add(
        sp_range,
        CellIsRule(operator="equal", formula=['"No"'],
                   fill=PatternFill(start_color="F8696B", end_color="F8696B", fill_type="solid")),
    )
    ws.conditional_formatting.add(
        sp_range,
        CellIsRule(operator="equal", formula=['"Yes"'],
                   fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")),
    )
    ws.conditional_formatting.add(
        sp_range,
        CellIsRule(operator="equal", formula=['"Likely Yes"'],
                   fill=PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")),
    )

    # Status dropdown (col 25)
    dv = DataValidation(
        type="list",
        formula1='"New,Shortlisted,Applied,Phone Screen,Interview,Offer,Rejected,Skipped"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)
    dv.add(f"Y2:Y{last_row}")

    # Referral dropdown (col 28)
    dv2 = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv2)
    dv2.add(f"AB2:AB{last_row}")

    wb.save(out_path)
    logger.info("xlsx written: %s (%d rows)", out_path, n_rows)
    return out_path
